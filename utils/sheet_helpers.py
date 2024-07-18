import os
from typing import Tuple

from openpyxl.styles import PatternFill

from bulk_upload.constants import ERROR_TEXT
from utils.constants import COMMENT_CELL_VALUE, ERROR_FILE
from copy import copy


def get_red_fill(fill_type: str = 'solid') -> PatternFill:
    return PatternFill(start_color=ERROR_TEXT, end_color=ERROR_TEXT, fill_type=fill_type)


def create_shell_with_style(sheet, cell_value: str, row: int, column: int, style,):
    comment_cell = sheet.cell(row=row, column=column)
    comment_cell._style = style
    comment_cell.value = cell_value
    return comment_cell


def check_sheet_titles(sheet, column_list: list, title_cell_style=None, comment_cell_value: str = COMMENT_CELL_VALUE
                       ) -> Tuple[bool, dict]:
    if not title_cell_style:
        title_cell_style = copy(sheet.cell(row=1, column=1)._style)
    index_sheet = {}
    sheet_errors = ""
    comment_column = 1
    status = True
    title_row = list(sheet.iter_rows(max_row=1, values_only=True))[0]
    for column, value in enumerate(title_row):
        if column > (len(column_list) - 1):
            break
        if not value:
            break
        comment_column += 1
        filed_name = value.lower().strip()
        if filed_name not in column_list:
            sheet_errors += f"Invalid column value '{value}' in column no. '{column + 1}'\n"
        else:
            index_sheet[filed_name] = column

    different_column = set(column_list).difference(set(list(index_sheet.keys())))
    if different_column:
        sheet_errors += f"Invalid or missing column {list(different_column)}\n"

    # If error then set error on comment column
    if sheet_errors:
        status = False
        create_shell_with_style(
            sheet=sheet, row=1, column=comment_column, cell_value=comment_cell_value, style=title_cell_style)
        sheet.cell(row=2, column=comment_column).value = sheet_errors

    return (
        status, {
            'sheet_indexes': index_sheet,
            "title_cell_style": title_cell_style,
            "error_column": comment_column-1,
            "errors": sheet_errors,
            "comment_column": comment_column
        })


def store_error_sheet(wb_obj, dir_path: str) -> str:
    abs_dir_path = os.path.abspath(dir_path)
    error_sheet_path = abs_dir_path + "/" + ERROR_FILE
    if not os.path.exists(abs_dir_path):
        os.makedirs(abs_dir_path)
    wb_obj.save(error_sheet_path)
    return error_sheet_path
